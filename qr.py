import inspect
import time
from io import BytesIO
import qrcode
from aiogram import Bot, Dispatcher, executor, types
import logging
from aiogram.dispatcher.filters import Text
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from transliterate import to_latin, to_cyrillic
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from bot_commands import set_default_commands
import environs
from openpyxl import Workbook

env = environs.Env()
env.read_env()
TOKEN = env.str("TOKEN")
ADMIN_ID = env.int("ADMIN_ID")

storage = MemoryStorage()
bot = Bot(token=TOKEN, parse_mode='HTML')
dp = Dispatcher(bot, storage=storage)

logging.basicConfig(format=u'%(filename)s [LINE:%(lineno)d] #%(levelname)-8s [%(asctime)s]  %(message)s', level=logging.INFO,)


class Cyrillic2latin(StatesGroup):
    cyrillic2latin = State()


class Qr(StatesGroup):
    qr = State()


class Ivmsfile(StatesGroup):
    ivmsfile = State()


class GetMyID(StatesGroup):
    getmyid = State()


def csv_to_xlsx(filename):
    possible_encodings = ['utf-8', 'windows-1251', 'iso-8859-5']
    for encoding in possible_encodings:
        try:
            with open(filename, 'r', encoding=encoding) as file:
                full_text = file.read()
                one_str = full_text.split('\n')

                data = []
                for i in one_str:
                    if i != '':
                        part_data = i.split(',')
                        data.append(part_data)

                people = {}
                for i in data:
                    if i[0] not in people.keys():
                        people[i[0]] = [[i[1], i[2], i[3]]]
                    else:
                        people[i[0]].append([i[1], i[2], i[3]])

                for i in people.keys():
                    sorted_data = []
                    for ii in people[i]:
                        if sorted_data == []:
                            if ii[2] == "Check-in":  # 1 эл
                                sorted_data.append(ii)
                        else:
                            if sorted_data[-1][2] != ii[2]:  # если checkin != checkin | т.е checkout
                                sorted_data.append(ii)  #
                            else:
                                sorted_data[-1] = ii  # если последний элемент == checkin, то перепишет значение

                    people[i] = sorted_data

                for i in people.keys():
                    data_struct = []

                    struct = {
                        'in': None,
                        'out': None,
                        'worked_hours': None,
                    }

                    for ii in people[i]:
                        if ii[2] == 'Check-in':
                            struct['in'] = [ii[0], ii[1]]
                        else:
                            struct['out'] = [ii[0], ii[1]]
                            data_struct.append(struct)
                            struct = {
                                'in': None,
                                'out': None,
                                'worked_hours': None,
                            }

                    people[i] = data_struct

                for i in people.keys():
                    for ii in people[i]:
                        ii['in'][0] = ii['in'][0].split('-')
                        ii['in'][1] = ii['in'][1].split(':')
                        ii['out'][0] = ii['out'][0].split('-')
                        ii['out'][1] = ii['out'][1].split(':')

                # calculating working hours
                for i in people.keys():
                    for ii in people[i]:
                        if ii['in'][0][0] == ii['out'][0][0]:
                            worktime = int(ii['out'][1][0]) - int(ii['in'][1][0])
                        else:
                            worktime = 24 - int(ii['in'][1][0]) + int(ii['out'][1][0])

                        if int(ii['in'][1][1]) > 15:
                            worktime -= 1

                        if int(ii['out'][1][1]) > 45:
                            worktime += 1

                        if worktime < 1:
                            worktime = 0

                        ii['worked_hours'] = worktime

            # Writing to xlxs
            wb = Workbook()
            ws = wb.active
            ws.title = 'Часы посещения'

            headers = ['Сотрудник', 'Дата', 'Время', 'Действие', 'Часы']
            ws.append(headers)

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", wrap_text=True)
            header_fill = PatternFill(start_color='abdbe3', end_color='abdbe3', fill_type='solid')
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                   bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
                cell.border = header_border

            # Iterate over all columns and adjust their widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 12

            for p in people:
                for ii in people[p]:
                    temp_date_in = '{0}-{1}-{2}'.format(ii['in'][0][0], ii['in'][0][1], ii['in'][0][2])
                    temp_date_out = '{0}-{1}-{2}'.format(ii['out'][0][0], ii['out'][0][1], ii['out'][0][2])

                    temp_time_in = '{0}:{1}'.format(ii['in'][1][0], ii['in'][1][1])
                    temp_time_out = '{0}:{1}'.format(ii['out'][1][0], ii['out'][1][1])

                    worked_hours = int(ii['worked_hours'])

                    first_row = [p, temp_date_in, temp_time_in, 'Пришел', '']
                    ws.append(first_row)

                    second_row = [p, temp_date_out, temp_time_out, 'Ушел', worked_hours]
                    ws.append(second_row)

                total_hours = 0
                for ii in people[p]:
                    total_hours += ii['worked_hours']

                total = ['', '', '', 'Всего:', total_hours]

                if total_hours == 0:
                    total = [i, '', '', 'Всего:', total_hours]

                ws.append(total)

                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)

                if len(p) > 1:
                    ws.append(['', '', '', '', ''])

            # Set cell styles
            for row_num in range(2, ws.max_row + 1):
                for col_num in range(1, ws.max_column + 1):
                    if col_num == 5:
                        ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center',
                                                                                          vertical='center')
                    else:
                        ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left')

                    cell_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    cell_fill = PatternFill(fill_type='solid', start_color='fcfcff', end_color='fcfcff') # light gray

                    ws.cell(row=row_num, column=col_num).border = cell_border
                    ws.cell(row=row_num, column=col_num).fill = cell_fill

            wb.save('results.xlsx')
            break
        except UnicodeDecodeError:
            logging.info('UnicodeDecodeError')
            pass


def generate_report(input_file, output_file):

    logging.info('Generating report...')
    wb_results = load_workbook(input_file)
    ws_results = wb_results.active

    wb_report = Workbook()
    ws_report = wb_report.active
    ws_report.title = 'Отчет о посещении'
    # ws_report.sheet_view.showGridLines = False

    headers = ['Сотрудник', 'План \nдни ', 'Факт \nдни', 'План \nчасы', 'Факт \nчасы',
               'Отраб. \nчасы в %', 'Комментарии', 'Снижение']

    ws_report.append(headers)

    # Iterate over all columns and adjust their widths
    for column in ws_report.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:

                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_report.column_dimensions[column_letter].width = adjusted_width

    for cell in ws_report[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color='DEE0EE', end_color='DEE0EE', fill_type='solid')  # gray
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        if cell.value == 'Снижение':
            cell.fill = PatternFill(start_color='Fbedca', end_color='Fbedca', fill_type='solid')  # orange

    ws_report.column_dimensions['A'].width = 35

    planned_work_days = 26
    planned_worked_hours = 225
    actual_work_days = 0
    actual_worked_hours = 0
    actual_worked_hours_in_percent = 0
    current_employee = None

    for row in ws_results.iter_rows(min_row=2, values_only=True):
        employee_name, date, time, action, work_hours = row

        # Check if a new employee has started
        if employee_name != current_employee:
            if current_employee is not None:
                actual_worked_hours_in_percent = (actual_worked_hours / planned_worked_hours) * 100 if actual_worked_hours != 0 else 0

                ws_report.append(
                    [
                        current_employee,
                        planned_work_days,
                        actual_work_days,
                        planned_worked_hours,
                        actual_worked_hours,
                        int(actual_worked_hours_in_percent),
                        '',
                        ''
                    ]
                )

                if actual_worked_hours_in_percent > 90:
                    ws_report.cell(row=ws_report.max_row, column=7).value = 'Отлично'
                    fill = PatternFill(start_color='C5f1ca', end_color='C5f1ca', fill_type='solid')  # light green
                    ws_report.cell(row=ws_report.max_row, column=6).fill = fill

                else:
                    fill = PatternFill(start_color='F7e2ee', end_color='F7e2ee', fill_type='solid')  # rose
                    ws_report.cell(row=ws_report.max_row, column=6).fill = fill

            current_employee = employee_name
            actual_work_days = 0
            actual_worked_hours = 0

        # Update actual work days and hours for the current employee
        if work_hours:
            actual_work_days += 1
            actual_worked_hours += work_hours

    # Append the report row for the last employee
    if current_employee is not None:
        actual_worked_hours_in_percent = (
                                                     actual_worked_hours / planned_worked_hours) * 100 if actual_worked_hours != 0 else 0

        ws_report.append(
            [
                current_employee,
                planned_work_days,
                actual_work_days,
                planned_worked_hours,
                actual_worked_hours,
                int(actual_worked_hours_in_percent),
                '',
                ''
            ]
        )

        if actual_worked_hours_in_percent > 90:
            ws_report.cell(row=ws_report.max_row, column=7).value = 'Отлично'
            fill = PatternFill(start_color='C5f1ca', end_color='C5f1ca', fill_type='solid')  # light green
            ws_report.cell(row=ws_report.max_row, column=7).fill = fill

        else:
            fill = PatternFill(start_color='F7e2ee', end_color='F7e2ee', fill_type='solid')  # rose
            ws_report.cell(row=ws_report.max_row, column=7).fill = fill

    # Set cell styles
    for row_num in range(2, ws_report.max_row + 1):
        for col_num in range(1, ws_report.max_column + 1):
            if col_num == 1:
                ws_report.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left',
                                                                                  vertical='center')
            else:
                ws_report.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

            cell_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                 top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            ws_report.cell(row=row_num, column=col_num).border = cell_border
            # cell_fill = PatternFill(fill_type='solid', start_color='fcfcff', end_color='fcfcff')  # light gray
            # ws_report.cell(row=row_num, column=col_num).fill = cell_fill

    wb_report.save(output_file)
    logging.info('Report generated successfully')


def auth(func):
    async def wrapper(*args, **kwargs):
        message = args[0]

        if 'state' in inspect.signature(func).parameters:
            return await func(*args, **kwargs)

        if message.from_user.id == ADMIN_ID:
            return await func(message)
        return await message.answer("Access denied!", reply=False)
    return wrapper


@dp.message_handler(commands=['start'])
@auth
async def cmd_start(message: types.Message):
    reply = f"Hello {message.from_user.full_name}! Choose any option from menu"
    logging.info("Bot started!")
    await message.answer(reply)


@dp.message_handler(state='*', commands='cancel')
@dp.message_handler(Text(equals='cancel', ignore_case=True), state='*')
@auth
async def cancel_handler(message: types.Message, state: FSMContext, **kwargs):
    current_state = await state.get_state()
    logging.info(f"current_state {current_state}")
    if current_state is None:
        return

    logging.info('Cancelling state %r', current_state)
    await state.finish()
    await message.reply(f'Cancelled <b>{current_state.split(":")[-1]}</b>')


@dp.message_handler(commands=['ivmsfile'], state=None)
@auth
async def ivmsfile(message: types.Message):
    logging.info(f"Starting {message.get_command()}")
    await message.answer(text="Send file from IVMS")
    await Ivmsfile.ivmsfile.set()


@dp.message_handler(content_types=['document'], state=Ivmsfile.ivmsfile)
async def file_echo(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        start_time = time.time()
        data['ivmsfile'] = message.document
        await message.answer('⏳ Wait a few seconds...')
        await data['ivmsfile'].download(destination_file='./download.csv')

        csv_to_xlsx('./download.csv')
        await message.reply_document(open('results.xlsx', 'rb'), caption='Ready!')

        await message.answer('⏳ Wait. Generating report')

        input_file = './results.xlsx'
        output_file = './report.xlsx'
        generate_report(input_file, output_file)
        await message.reply_document(open('./report.xlsx', 'rb'), caption='Report is ready!')

        end_time = time.time()
        elapsed_time = end_time - start_time

        if elapsed_time > 60:
            await message.answer(f"Time used: {elapsed_time / 60:.2f} min")
        else:
            await message.answer(f"Time used: {elapsed_time:.2f} sec")
        logging.info("Files done")


@dp.message_handler(commands=['cyrillic2latin'], state=None)
@auth
async def cyrillic2latin(message: types.Message):
    await message.answer(text="Insert text")
    await Cyrillic2latin.cyrillic2latin.set()


@dp.message_handler(state=Cyrillic2latin.cyrillic2latin)
async def cyrillic2latin_handler(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['cyrillic2latin'] = message.text

        if data['cyrillic2latin'].isascii():
            reply = to_cyrillic(data['cyrillic2latin'])
        else:
            reply = to_latin(data['cyrillic2latin'])
        if data['cyrillic2latin'].isnumeric():
            reply = "Insert text only"

        await message.answer(reply)


@dp.message_handler(commands=['qr'], state=None)
@auth
async def qr(message: types.Message):
    await message.answer(text="Insert url")
    await Qr.qr.set()


@dp.message_handler(state=Qr.qr)
async def qr_handler(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['qr'] = message.text
        qr_img = qrcode.make(data['qr'])
        bio = BytesIO()
        qr_img.save(bio, 'JPEG')
        bio.seek(0)
        await message.answer_photo(bio, caption=f"QR code of {data['qr']}")


@dp.message_handler(commands=['getmyid'], content_types=['text'], state=None)
@auth
async def get_my_id(message: types.Message):
    await message.answer(f"<b>Your ID:</b> {message.from_user.id}")
    await message.answer(text="If you want to get someone else's ID forward message")
    await GetMyID.getmyid.set()


@dp.message_handler(state=GetMyID.getmyid)
async def get_my_id_handler(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['getmyid'] = message.text
        current_user_id = message.from_user.id
        current_chat_id = message.chat.id

        if message.forward_from:
            await message.answer(f"<b>Ваш ID:</b> {current_user_id}\n<b>Current chat ID:</b> {current_chat_id}\n<b>Forwarded from ID:</b> {message.forward_from.id}")
        else:
            await message.reply(f'🚫 Unsupported format')
            await message.delete()


async def on_startup(dispatcher):
    await set_default_commands(dispatcher)


executor.start_polling(dp, on_startup=on_startup, skip_updates=False)
